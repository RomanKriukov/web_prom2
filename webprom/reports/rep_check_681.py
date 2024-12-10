import os
import pandas as pd
import numpy as np
from sqlalchemy import text
from sqlalchemy.exc import DBAPIError
from sqlalchemy.orm import Session
from flask import current_app
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment


def rep_check_681(session:Session, date_since: str, date_till: str, path: str, account: int) -> str:
    try:
        query = f"EXEC repCheck681 '{date_since.strftime("%Y%m%d")}', '{date_till.strftime("%Y%m%d")}', '{path}', {account}"
        rs = session.execute(text(query)).fetchall()
        df = pd.DataFrame(rs)
    except DBAPIError as e:
        return e._sql_message()
    except Exception as e:
        return e 
    finally:
        session.close()

    df = df.astype({
        'summaNDS': 'float64',
        'osv681': 'float64',
        'diff': 'float64'
    })

    df.replace(np.nan, '', inplace=True)

    columns_to_show = ['firm', 'firmCode', 'summaNDS', 'osv681', 'diff']
    columns_name = ['Наименование', 'ЕДРПОУ', 'Сумма 1С', 'Сумма ФА', 'Отклонение']
    columns_width = [60, 10, 10, 10, 10]
    name = dict(zip(columns_to_show, columns_name))
    width = dict(zip(columns_to_show, columns_width))

    # HTML таблица
    html = '<table class="treetable" width="98%">\n'

    html += '<thead><tr>'
    for col in df.columns:       
        html += f'<th style="width:{width[col]}%">{name[col]}</th>' if col in columns_to_show else ''
    html += '</tr></thead><tbody>\n'

    for index, row in df.iterrows():
        html += f'<tr class="lev{row["levl"]}">'                 
        for col in df.columns: 
            if col in columns_to_show:
                if col == 'firm' and row["levl"] < 2 :
                    html += f'<td><label><input type="checkbox"><a onclick="sh(this)">{row[col]}</a></label></td>'
                else:
                    html += f'<td>{int(row[col]):,}</td>'.replace(",", " ") if type(row[col]) == float else f'<td>{row[col]}</td>'
        html += '</tr>\n'

    html += '</tbody></table>'
    
    # Excel файл
    output_file = 'Check681.xlsx'
    filepath = os.path.join(current_app.config['UPLOAD_FOLDER'], output_file) 
    if os.path.isfile(filepath):
        os.remove(filepath)
    try:
        wb = Workbook()
        ws = wb.active
        ws.title = "Check681"

        # оставляем только нужные заголовки (к тем, что отображаются в html добавляем levl)
        df = df[columns_to_show+['levl']]
        # переименовываем заголовки
        df.columns = columns_name+['levl']
        # записываем заголовки
        ws.append(df.columns.tolist())

        # построчно переносим фрейм в файл
        for r in dataframe_to_rows(df, index=False, header=False):
            ws.append(r)

        # форматируем заголовки
        for cell in ws[1]:
            cell.font = cell.font = Font(italic=True, bold=True, color='FF0000')
            cell.alignment = Alignment(horizontal='center', vertical='center')

        # группируем

        # это версия AI
        # row_start = None
        # for idx, levl in enumerate(df['levl']):
        #     if levl == 1:
        #         if row_start is not None and row_start < idx:
        #             ws.row_dimensions[row_start].outline_level = 1
        #             ws.row_dimensions[row_start].collapsed = False
        #             ws.row_dimensions[idx].outline_level = 1
        #             ws.row_dimensions[idx].collapsed = False
        #         row_start = idx + 2
        #     else:
        #         if row_start is not None and row_start < idx:
        #             ws.row_dimensions[row_start].outline_level = 1
        #             ws.row_dimensions[idx].outline_level = 1

        # это моя версия
        # странная группировка у экселя
        # 1. collapsed - заголовок группы
        # 2. от collapsed до collapsed с определенным outline_level - элементы группы       
        # 3. idx = 0 => A2 => A idx+2 
        # 4. collapsed = True - у меня не сработала
        # группировал как глав.бух и плановик: 1я строка - итог по 1й группе и не разворачивается, итог по 2й группе разворачивает первую группу и т.д.
        for idx, levl in enumerate(df['levl']):
            if levl == 1:
                ws.row_dimensions[idx+2].collapsed = False
                # форматируем заголовки групп
                for cell in ws[idx+2]:
                    cell.font = Font(bold=True)
                    cell.fill = PatternFill(start_color='E7E6E6', end_color='E7E6E6', fill_type='solid')
                    cell.border = Border(
                        left=Side(border_style='thin', color='000000'),
                        right=Side(border_style='thin', color='000000'),
                        top=Side(border_style='thin', color='000000'),
                        bottom=Side(border_style='thin', color='000000')
                    )
            else:
                ws.row_dimensions[idx+2].outline_level = 1
        
        # удаляем столбец levl
        ws.delete_cols(6)

        for i, w in enumerate(columns_width):
            ws.column_dimensions[chr(65+i)].width = w + 5   # ASCII code 'A' = 65

        wb.save(filepath)
    except Exception as e:
        #current_app.logger.info(os.system('whoami'))
        current_app.logger.info(e)
        return [html, None]        
    else:
        return [html, output_file]